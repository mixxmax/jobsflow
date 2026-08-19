#!/usr/bin/env python3
"""Sync material status in the 2026-07-26 CSV/XLSX trackers based on validated packages."""
import argparse
import csv
import json
import sys
from pathlib import Path

from openpyxl import load_workbook


def sync_tracker_status(manifest_path: Path, csv_path: Path, xlsx_path: Path) -> dict:
    """Update material status for validated packages. Returns dict keyed by URL."""
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    viable_urls = {j["url"] for j in manifest["viable"]}
    excluded_urls = {j["url"] for j in manifest["excluded_f"]}

    # Determine which viable packages pass validation
    from tools.core_applications.validate_package import validate_package
    passed = {}
    blocked = {}
    for j in manifest["viable"]:
        d = Path(j["parent_dir"]) / j["folder_name"]
        errs = validate_package(d, j["company"], j["role"])
        if errs:
            blocked[j["url"]] = errs
        else:
            passed[j["url"]] = "已制作"

    # Update CSV
    rows = []
    with open(csv_path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames
        for row in reader:
            url = row.get("链接", "").strip()
            if url in passed:
                row["材料状态"] = "已制作"
            # A blocked package stays at its existing status; ``待核实`` is
            # not a valid V-column option and would violate the tracker
            # contract.  F-grade exclusions likewise remain unchanged.
            rows.append(row)

    with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    # Update XLSX (all sheets)
    wb = load_workbook(xlsx_path)
    for ws in wb.worksheets:
        # Find URL and status columns
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        url_col = None
        status_col = None
        for i, h in enumerate(header, 1):
            if h and "链接" in str(h):
                url_col = i
            if h and "材料状态" in str(h):
                status_col = i
        if not url_col or not status_col:
            continue
        for row in ws.iter_rows(min_row=2):
            cell_url = row[url_col - 1].value
            if cell_url:
                url = str(cell_url).strip()
                if url in passed:
                    row[status_col - 1].value = "已制作"

    wb.save(xlsx_path)

    return {
        "passed": len(passed),
        "blocked": len(blocked),
        "total_viable": len(manifest["viable"]),
        "total_excluded": len(manifest["excluded_f"]),
    }


def main():
    parser = argparse.ArgumentParser(description="Sync tracker material status")
    parser.add_argument("--manifest", required=True)
    parser.add_argument("--csv", required=True)
    parser.add_argument("--xlsx", required=True)
    args = parser.parse_args()

    result = sync_tracker_status(
        Path(args.manifest), Path(args.csv), Path(args.xlsx)
    )
    print(f"Passed: {result['passed']}, Blocked: {result['blocked']}, "
          f"Total viable: {result['total_viable']}, Excluded: {result['total_excluded']}")


if __name__ == "__main__":
    main()
